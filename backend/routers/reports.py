"""
routers/reports.py — Report generation endpoints (Excel + PDF)

Endpoints:
    GET /societies/{society_id}/reports/collection    ?month=&year=  → xlsx / pdf
    GET /societies/{society_id}/reports/defaulters    ?month=&year=  → xlsx / pdf
    GET /societies/{society_id}/reports/expenses      ?year=         → xlsx / pdf
    GET /societies/{society_id}/reports/budget        ?fy=           → xlsx / pdf
    GET /societies/{society_id}/reports/visitors      ?date=         → xlsx / pdf
    GET /societies/{society_id}/reports/assets                       → xlsx / pdf
    GET /societies/{society_id}/reports/audit         ?limit=        → xlsx / pdf
"""

import io
from datetime import date, datetime
from typing import Literal, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from database import get_db
from auth import get_current_active_user, require_admin
from models import (
    Bill, Transaction, Visitor, Asset, AuditLog, Budget,
    User, Society
)

router = APIRouter(prefix="/societies/{society_id}/reports", tags=["reports"])


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _require_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")


def _require_reportlab():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        return SimpleDocTemplate, Table, TableStyle, Paragraph, getSampleStyleSheet, A4, colors
    except ImportError:
        raise HTTPException(500, "reportlab not installed. Run: pip install reportlab")


def _xlsx_response(wb, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


def _pdf_response(buf: io.BytesIO, filename: str) -> StreamingResponse:
    buf.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(buf, media_type="application/pdf", headers=headers)


def _xlsx_header_style(ws, openpyxl, row: int):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor="1E3A5F")
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


# ──────────────────────────────────────────────────────────────────────────────
# Collection Report
# ──────────────────────────────────────────────────────────────────────────────

@router.get("/collection")
def collection_report(
    society_id: int,
    month: Optional[int] = None,
    year: Optional[int] = None,
    fmt: Literal["xlsx", "pdf"] = "xlsx",
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    q = db.query(Bill).filter(Bill.society_id == society_id)
    if month:
        q = q.filter(Bill.bill_month == month)
    if year:
        q = q.filter(Bill.bill_year == year)
    bills = q.all()

    title = f"Collection Report — {month or 'All'}/{year or 'All'}"

    if fmt == "xlsx":
        openpyxl = _require_openpyxl()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Collection"
        headers = ["Flat No", "Month", "Year", "Amount (₹)", "Paid (₹)", "Balance (₹)", "Status", "Due Date"]
        ws.append(headers)
        _xlsx_header_style(ws, openpyxl, 1)
        for b in bills:
            ws.append([
                b.flat.flat_no if b.flat else "-",
                b.bill_month, b.bill_year,
                float(b.total_amount), float(b.paid_amount),
                float(b.total_amount - b.paid_amount),
                b.status,
                b.due_date.isoformat() if b.due_date else "-",
            ])
        # Summary row
        ws.append([])
        ws.append(["TOTAL", "", "", sum(float(b.total_amount) for b in bills),
                   sum(float(b.paid_amount) for b in bills), "", "", ""])
        return _xlsx_response(wb, f"collection_{month}_{year}.xlsx")

    # PDF
    SimpleDocTemplate, Table, TableStyle, Paragraph, getSampleStyleSheet, A4, colors = _require_reportlab()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    data = [["Flat", "Month/Year", "Amount", "Paid", "Balance", "Status"]]
    for b in bills:
        data.append([
            b.flat.flat_no if b.flat else "-",
            f"{b.bill_month}/{b.bill_year}",
            f"₹{b.total_amount:,.0f}",
            f"₹{b.paid_amount:,.0f}",
            f"₹{b.total_amount - b.paid_amount:,.0f}",
            b.status,
        ])
    t = Table(data)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#1A1F2C"), colors.HexColor("#13161E")]),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#E8EAF0")),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#2A3047")),
    ]))
    doc.build([Paragraph(title, styles["Title"]), t])
    return _pdf_response(buf, f"collection_{month}_{year}.pdf")


# ──────────────────────────────────────────────────────────────────────────────
# Defaulters Report
# ──────────────────────────────────────────────────────────────────────────────

@router.get("/defaulters")
def defaulters_report(
    society_id: int,
    month: Optional[int] = None,
    year: Optional[int] = None,
    fmt: Literal["xlsx", "pdf"] = "xlsx",
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    q = db.query(Bill).filter(
        Bill.society_id == society_id,
        Bill.status.in_(["pending", "overdue", "partial"])
    )
    if month:
        q = q.filter(Bill.bill_month == month)
    if year:
        q = q.filter(Bill.bill_year == year)
    defaulters = q.order_by(Bill.due_date).all()

    if fmt == "xlsx":
        openpyxl = _require_openpyxl()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Defaulters"
        ws.append(["Flat", "Resident", "Month/Year", "Due Amount (₹)", "Status", "Due Date", "Days Overdue"])
        _xlsx_header_style(ws, openpyxl, 1)
        today = date.today()
        for b in defaulters:
            overdue_days = (today - b.due_date).days if b.due_date and b.due_date < today else 0
            resident_name = b.flat.users[0].full_name if b.flat and b.flat.users else "-"
            ws.append([
                b.flat.flat_no if b.flat else "-",
                resident_name,
                f"{b.bill_month}/{b.bill_year}",
                float(b.total_amount - b.paid_amount),
                b.status,
                b.due_date.isoformat() if b.due_date else "-",
                overdue_days,
            ])
        return _xlsx_response(wb, f"defaulters_{month}_{year}.xlsx")

    # PDF
    SimpleDocTemplate, Table, TableStyle, Paragraph, getSampleStyleSheet, A4, colors = _require_reportlab()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    data = [["Flat", "Month/Year", "Balance Due", "Status"]]
    today = date.today()
    for b in defaulters:
        data.append([
            b.flat.flat_no if b.flat else "-",
            f"{b.bill_month}/{b.bill_year}",
            f"₹{b.total_amount - b.paid_amount:,.0f}",
            b.status,
        ])
    t = Table(data)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#450A0A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#2A3047")),
    ]))
    doc.build([Paragraph(f"Defaulters Report — {month or 'All'}/{year or 'All'}", styles["Title"]), t])
    return _pdf_response(buf, f"defaulters_{month}_{year}.pdf")


# ──────────────────────────────────────────────────────────────────────────────
# Expense Report
# ──────────────────────────────────────────────────────────────────────────────

@router.get("/expenses")
def expense_report(
    society_id: int,
    year: Optional[int] = None,
    fmt: Literal["xlsx", "pdf"] = "xlsx",
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    q = db.query(Transaction).filter(
        Transaction.society_id == society_id,
        Transaction.transaction_type == "expense"
    )
    if year:
        q = q.filter(Transaction.transaction_date.between(
            datetime(year, 1, 1), datetime(year, 12, 31, 23, 59, 59)
        ))
    txns = q.order_by(Transaction.transaction_date.desc()).all()

    if fmt == "xlsx":
        openpyxl = _require_openpyxl()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expenses"
        ws.append(["Date", "Category", "Description", "Amount (₹)", "Payment Mode", "Invoice"])
        _xlsx_header_style(ws, openpyxl, 1)
        for t in txns:
            ws.append([
                t.transaction_date.date().isoformat() if t.transaction_date else "-",
                t.category,
                t.description,
                float(t.amount),
                t.payment_mode or "-",
                t.invoice_url or "-",
            ])
        ws.append([])
        ws.append(["TOTAL", "", "", sum(float(t.amount) for t in txns)])
        return _xlsx_response(wb, f"expenses_{year}.xlsx")

    SimpleDocTemplate, Table, TableStyle, Paragraph, getSampleStyleSheet, A4, colors = _require_reportlab()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    data = [["Date", "Category", "Description", "Amount"]]
    for t in txns:
        data.append([
            t.transaction_date.date().isoformat() if t.transaction_date else "-",
            t.category, t.description or "-", f"₹{t.amount:,.0f}",
        ])
    tbl = Table(data, colWidths=[80, 80, 200, 80])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#431407")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#2A3047")),
    ]))
    doc.build([Paragraph(f"Expense Report — FY {year}", styles["Title"]), tbl])
    return _pdf_response(buf, f"expenses_{year}.pdf")


# ──────────────────────────────────────────────────────────────────────────────
# Asset Register
# ──────────────────────────────────────────────────────────────────────────────

@router.get("/assets")
def asset_report(
    society_id: int,
    fmt: Literal["xlsx", "pdf"] = "xlsx",
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    assets = db.query(Asset).filter(Asset.society_id == society_id).all()

    if fmt == "xlsx":
        openpyxl = _require_openpyxl()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Assets"
        ws.append(["Asset Name", "Category", "Status", "Purchase Value (₹)",
                   "Last Service", "Next Service", "Location"])
        _xlsx_header_style(ws, openpyxl, 1)
        for a in assets:
            ws.append([
                a.name, a.category, a.status,
                float(a.purchase_value) if a.purchase_value else 0,
                a.last_service_date.isoformat() if a.last_service_date else "-",
                a.next_service_date.isoformat() if a.next_service_date else "-",
                a.location or "-",
            ])
        return _xlsx_response(wb, "asset_register.xlsx")

    SimpleDocTemplate, Table, TableStyle, Paragraph, getSampleStyleSheet, A4, colors = _require_reportlab()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    data = [["Asset", "Category", "Status", "Value", "Next Service"]]
    for a in assets:
        data.append([
            a.name, a.category, a.status,
            f"₹{a.purchase_value:,.0f}" if a.purchase_value else "-",
            a.next_service_date.isoformat() if a.next_service_date else "-",
        ])
    tbl = Table(data)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E3A2F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#2A3047")),
    ]))
    doc.build([Paragraph("Asset Register", styles["Title"]), tbl])
    return _pdf_response(buf, "asset_register.pdf")


# ──────────────────────────────────────────────────────────────────────────────
# Visitor Log
# ──────────────────────────────────────────────────────────────────────────────

@router.get("/visitors")
def visitor_report(
    society_id: int,
    visit_date: Optional[date] = Query(None, alias="date"),
    fmt: Literal["xlsx", "pdf"] = "xlsx",
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    q = db.query(Visitor).filter(Visitor.society_id == society_id)
    if visit_date:
        q = q.filter(Visitor.created_at >= datetime.combine(visit_date, datetime.min.time()),
                     Visitor.created_at < datetime.combine(visit_date, datetime.max.time()))
    visitors = q.order_by(Visitor.check_in_time.desc()).all()

    if fmt == "xlsx":
        openpyxl = _require_openpyxl()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visitors"
        ws.append(["Visitor Name", "Mobile", "Purpose", "Flat", "Status",
                   "Check-In", "Check-Out", "OTP"])
        _xlsx_header_style(ws, openpyxl, 1)
        for v in visitors:
            ws.append([
                v.visitor_name, v.visitor_mobile or "-", v.purpose or "-",
                v.flat.flat_no if v.flat else "-",
                v.status,
                str(v.check_in_time) if v.check_in_time else "-",
                str(v.check_out_time) if v.check_out_time else "-",
                v.otp_code or "-",
            ])
        return _xlsx_response(wb, f"visitors_{visit_date or 'all'}.xlsx")

    SimpleDocTemplate, Table, TableStyle, Paragraph, getSampleStyleSheet, A4, colors = _require_reportlab()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    data = [["Name", "Purpose", "Flat", "Status", "Check-In"]]
    for v in visitors:
        data.append([
            v.visitor_name, v.purpose or "-",
            v.flat.flat_no if v.flat else "-",
            v.status,
            str(v.check_in_time)[:16] if v.check_in_time else "-",
        ])
    tbl = Table(data)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#422006")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#2A3047")),
    ]))
    doc.build([Paragraph(f"Visitor Log — {visit_date or 'All dates'}", styles["Title"]), tbl])
    return _pdf_response(buf, f"visitors_{visit_date or 'all'}.pdf")
